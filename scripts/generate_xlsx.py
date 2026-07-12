#!/usr/bin/env python3
"""
generate_xlsx.py — Generates Cloud_Intelligence_Matrix.xlsx from data/matrix.json
Reads new capability-v1 schema.
Output: dist/Cloud_Intelligence_Matrix.xlsx plus dated per-view CSV files
"""
import csv
import json, sys
from datetime import date
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

ROOT   = Path(__file__).parent.parent
DATA   = ROOT / "data"
OUTDIR = ROOT / "dist"
OUTDIR.mkdir(exist_ok=True)

mdata = json.loads((DATA / "matrix.json").read_text())
udata = json.loads((DATA / "upcoming.json").read_text())
hdata = json.loads((DATA / "history.json").read_text())
tdata = json.loads((DATA / "transparency.json").read_text())
fdata = json.loads((DATA / "federal_transparency.json").read_text())
idata = json.loads((DATA / "international_transparency.json").read_text())
sdata = json.loads((DATA / "status.json").read_text())
adata = json.loads((DATA / "ai_watch.json").read_text())

META       = mdata["_meta"]
CAPS       = mdata["capabilities"]
CATEGORIES = mdata["categories"]
TAG_DEFS   = mdata["tags"]
FRAMEWORKS = mdata["frameworks"]
CONTROL_LENS = mdata["controlLens"]
COMPLIANCE_FRAMEWORKS = mdata.get("complianceFrameworks", [])
PATTERNS   = mdata["patterns"]
TIERS      = META["tiers"]
PROVIDERS  = META["providers"]
UPCOMING   = udata.get("upcoming", [])
HISTORY    = hdata.get("history", [])
TRANSPARENCY = tdata.get("mandates", [])
TRANSPARENCY_META = tdata.get("_meta", {})
FEDERAL_TRANSPARENCY = fdata.get("records", [])
FEDERAL_TRANSPARENCY_META = fdata.get("_meta", {})
INTERNATIONAL_TRANSPARENCY = idata.get("records", [])
INTERNATIONAL_TRANSPARENCY_META = idata.get("_meta", {})
STATUS_SOURCES = sdata.get("sources", [])
AI_WATCH_SOURCES = adata.get("sources", [])
CAP_MAP    = {cap["capability"]: cap for cap in CAPS}

PROV_LABELS = {"aws":"AWS","azure":"Azure","gcp":"GCP","oci":"OCI"}
EXPORT_DATE = date.today().isoformat()
PROV_COLORS = {
    "aws":   {"hdr":"B45309","note":"FEF3C7","svc":"FFFBEB"},
    "azure": {"hdr":"0E7490","note":"CFFAFE","svc":"ECFEFF"},
    "gcp":   {"hdr":"1A56A0","note":"DBEAFE","svc":"EFF6FF"},
    "oci":   {"hdr":"9F3727","note":"FCE7E4","svc":"FFF4F2"},
}
MATRIX_COLS = 4 + len(PROVIDERS) * 2
TAG_FG = {
    "gray":"6B7280","blue":"2563EB","purple":"7C3AED","green":"15803D",
    "amber":"B45309","red":"B91C1C","cyan":"0891B2","slate":"475569",
    "teal":"0D9488","orange":"C2410C","yellow":"A16207","rose":"BE123C",
}
GOV_COLORS = {
    "Full":"065F46","Partial":"78350F","Limited":"7C2D12","None":"374151",
}
PARITY_COLORS = {
    "None":"374151","Minor":"92400E","Moderate":"B45309","Significant":"991B1B",
}

THIN = Side(style="thin", color="D0D7E3")
TB   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def f(h):     return PatternFill("solid", fgColor=h.lstrip("#"))
def ft(bold=False, size=9, color="111827", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color.lstrip("#"), italic=italic)
def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws["A1"]
    c.value = text
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = f("0F1A2E")
    c.alignment = al("center","center")
    ws.row_dimensions[1].height = 26

def build_matrix_sheet(ws, tier=None):
    """One sheet: capability rows, provider cols (service + gov + parity), optionally filtered to tier."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    title = f"Cloud Intelligence Matrix — {tier or 'All Tiers'}"
    hdr(ws, title, MATRIX_COLS)

    # Row 2: column headers
    col_headers = [
        ("Category",    "1E3A5F"),
        ("Capability",  "1E3A5F"),
        ("Tags",        "1E3A5F"),
    ]
    for pkey in PROVIDERS:
        col_headers.extend([
            (f"{PROV_LABELS[pkey]} Service", PROV_COLORS[pkey]["hdr"]),
            (f"{PROV_LABELS[pkey]} Gov", PROV_COLORS[pkey]["hdr"]),
        ])
    col_headers.append(("Verified", "1E3A5F"))
    for ci, (h, bg) in enumerate(col_headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = f(bg); c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center","center"); c.border = TB
    ws.row_dimensions[2].height = 16

    row = 3
    last_cat = None
    for cap in CAPS:
        cat = cap["category"]
        bg_cat = "F0F9FF" if CATEGORIES.index(cat) % 2 == 0 else "F8FAFC"

        # Category cell
        cc = ws.cell(row=row, column=1, value=cap["category"] if cat != last_cat else "")
        cc.fill = f("E0F2FE" if cat != last_cat else bg_cat)
        cc.font = Font(name="Arial", bold=True, size=8, color="0369A1")
        cc.alignment = al("center","center", wrap=True); cc.border = TB
        last_cat = cat

        # Capability
        capcel = ws.cell(row=row, column=2, value=cap["capability"])
        capcel.fill = f(bg_cat); capcel.font = ft(bold=True, size=9, color="111827")
        capcel.alignment = al("left","center", wrap=True); capcel.border = TB

        # Tags
        tc = ws.cell(row=row, column=3, value=" · ".join(cap.get("tags",[])))
        tc.fill = f(bg_cat); tc.font = ft(size=8, italic=True, color="374151")
        tc.alignment = al("left","center", wrap=True); tc.border = TB

        # Provider columns (svc + gov)
        for pi, pkey in enumerate(PROVIDERS):
            prov = cap.get("providers",{}).get(pkey,{})
            svc  = prov.get("service","—")
            if tier:
                svc = prov.get("tierNotes",{}).get(tier, svc)
            gov  = prov.get("govAvailability","—")
            par  = prov.get("parityLag","None")
            pc   = PROV_COLORS[pkey]

            # Service
            sc = ws.cell(row=row, column=4+pi*2, value=svc)
            sc.fill = f(pc["svc"]); sc.font = ft(size=9)
            sc.alignment = al("left","center",wrap=True); sc.border = TB

            # Gov + parity
            gov_txt = gov
            if par and par != "None": gov_txt += f" · LAG:{par}"
            gc = ws.cell(row=row, column=5+pi*2, value=gov_txt)
            gov_bg = {"Full":"D1FAE5","Partial":"FEF3C7","Limited":"FFE4E6","None":"F9FAFB"}.get(gov,"F9FAFB")
            gc.fill = f(gov_bg); gc.font = ft(size=8, bold=(gov!="Full"), color=GOV_COLORS.get(gov,"374151"))
            gc.alignment = al("center","center"); gc.border = TB

        # Verified
        vc = ws.cell(row=row, column=MATRIX_COLS, value=cap.get("lastVerified",""))
        vc.fill = f(bg_cat); vc.font = ft(size=8, color="6B7280")
        vc.alignment = al("center","center"); vc.border = TB

        ws.row_dimensions[row].height = 34
        row += 1

    # Widths
    for col, w in {"A":22,"B":30,"C":36}.items():
        ws.column_dimensions[col].width = w
    for pi, _ in enumerate(PROVIDERS):
        ws.column_dimensions[get_column_letter(4 + pi * 2)].width = 38
        ws.column_dimensions[get_column_letter(5 + pi * 2)].width = 18
    ws.column_dimensions[get_column_letter(MATRIX_COLS)].width = 12

def build_detail_sheet(ws):
    """Full detail: one row per capability×provider with arch notes and operational considerations."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    hdrs = ["Category","Capability","Provider","Service","Former Names","Gov Avail","Parity Lag","Parity Detail","Gov Variant",
            "Region","Realm Class","Provider Verified","Constraints","Cost Shape","Egress Sensitive","Commitment Discount",
            "PQC Status","PQC KEM","PQC Signature","PQC TLS","PQC VPN","PQC Milestone","PQC FIPS Parity","PQC Gov","PQC Source",
            "PQC Source Date","PQC First Party","PQC Confidence","PQC Note","Residency Offerings","Residency Geographies",
            "Residency Statuses","Residency Partner Operated","Residency Sources","FedRAMP Commercial Status","FedRAMP Commercial URL",
            "FedRAMP Commercial Date","FedRAMP Commercial Confidence","FedRAMP Government Status","FedRAMP Government DoD IL",
            "FedRAMP Government Boundary","FedRAMP Government URL","FedRAMP Government Date","FedRAMP Government Confidence",
            "FedRAMP Level","DoD Impact Level","Docs","Pricing","Compliance","Architecture Notes","Operational Considerations"]
    hdr(ws, "Capability Detail - Architecture Notes - Operational Considerations - All Tier Notes", len(hdrs))
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = f("1E3A5F"); c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center","center"); c.border = TB
    ws.row_dimensions[2].height = 16

    row = 3
    for cap in CAPS:
        for pi, pkey in enumerate(PROVIDERS):
            prov = cap.get("providers",{}).get(pkey,{})
            cost_model = prov.get("costModel", {})
            pqc = prov.get("pqcReadiness", {})
            residency = prov.get("residency", [])
            fedramp = prov.get("fedramp", {})
            fedramp_commercial = field_value(fedramp, "commercial") if isinstance(field_value(fedramp, "commercial"), dict) else {}
            fedramp_government = field_value(fedramp, "government") if isinstance(field_value(fedramp, "government"), dict) else {}
            bg = PROV_COLORS[pkey]["svc"]
            vals = [
                cap["category"], cap["capability"], PROV_LABELS[pkey],
                prov.get("service",""), cell_text(prov.get("formerNames","")), prov.get("govAvailability",""), prov.get("parityLag",""),
                prov.get("parityDetail",""), prov.get("govVariant",""), prov.get("region",""), prov.get("realmClass",""), prov.get("lastVerified",""),
                cell_text(prov.get("constraints","")), field_value(cost_model, "shape"), field_value(cost_model, "egressSensitive"),
                field_value(cost_model, "commitmentDiscountAvailable"), field_value(pqc, "status"), field_value(pqc, "kem"),
                field_value(pqc, "signature"), field_value(pqc, "tls"), field_value(pqc, "vpn"), field_value(pqc, "milestoneDate"),
                field_value(pqc, "fipsEndpointParity"), field_value(pqc, "govPqc"), field_value(pqc, "source"), field_value(pqc, "sourceDate"),
                field_value(pqc, "firstParty"), field_value(pqc, "confidence"), field_value(pqc, "note"),
                residency_field(residency, "offering"), residency_field(residency, "geography"), residency_field(residency, "status"),
                residency_partner_operated(residency), residency_field(residency, "source"),
                field_value(fedramp_commercial, "status"), field_value(fedramp_commercial, "url"), field_value(fedramp_commercial, "date"),
                field_value(fedramp_commercial, "confidence"), field_value(fedramp_government, "status"), field_value(fedramp_government, "dodIL"),
                field_value(fedramp_government, "boundary"), field_value(fedramp_government, "url"), field_value(fedramp_government, "date"),
                field_value(fedramp_government, "confidence"), prov.get("fedrampLevel",""), prov.get("dodImpactLevel",""),
                prov.get("docsUrl",""), prov.get("pricingUrl",""), prov.get("complianceUrl",""),
                cap.get("architectureNotes",""), cap.get("operationalConsiderations",""),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = f(bg); c.font = ft(size=8)
                c.alignment = al("left","center", wrap=True); c.border = TB
            ws.row_dimensions[row].height = 40
            row += 1

    for col, w in {
        "A":16,"B":22,"C":8,"D":30,"E":32,"F":12,"G":14,"H":34,"I":26,"J":24,"K":16,"L":14,
        "M":36,"N":16,"O":14,"P":18,"Q":16,"R":18,"S":18,"T":18,"U":18,"V":18,"W":18,"X":42,
        "Y":16,"Z":16,"AA":44,"AB":44,"AC":44,"AD":60,"AE":60,
    }.items():
        ws.column_dimensions[col].width = w

def build_gov_sheet(ws):
    """Government and parity focus sheet."""
    ws.sheet_view.showGridLines = False
    hdr(ws, "Government / Sovereign Cloud Availability & Parity Lag — Cloud Intelligence Matrix", MATRIX_COLS)
    hdrs = ["Category","Capability","Tags"]
    for pkey in PROVIDERS:
        hdrs.extend([f"{PROV_LABELS[pkey]} Gov", f"{PROV_LABELS[pkey]} Parity"])
    hdrs.append("Verified")
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = f("7F1D1D"); c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center","center"); c.border = TB
    ws.row_dimensions[2].height = 16

    row = 3
    for cap in CAPS:
        has_issue = any(
            cap["providers"].get(p,{}).get("govAvailability") != "Full" or
            (cap["providers"].get(p,{}).get("parityLag","None") != "None")
            for p in PROVIDERS
        )
        bg = "FFF1F2" if has_issue else "F9FAFB"
        vals_head = [cap["category"], cap["capability"], " · ".join(cap.get("tags",[]))]
        for ci, v in enumerate(vals_head, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = f(bg); c.font = ft(bold=(ci==2), size=9)
            c.alignment = al("left","center", wrap=True); c.border = TB

        for pi, pkey in enumerate(PROVIDERS):
            prov = cap.get("providers",{}).get(pkey,{})
            gov = prov.get("govAvailability","—")
            par = prov.get("parityLag","—")
            gov_bg = {"Full":"D1FAE5","Partial":"FEF3C7","Limited":"FFE4E6","None":"F9FAFB"}.get(gov,"F9FAFB")
            par_bg = {"None":"F9FAFB","Minor":"FEF3C7","Moderate":"FFE4E6","Significant":"FEE2E2"}.get(par,"F9FAFB")
            gc = ws.cell(row=row, column=4+pi*2, value=gov)
            gc.fill = f(gov_bg); gc.font = ft(bold=(gov!="Full"), size=8)
            gc.alignment = al("center","center"); gc.border = TB
            pc = ws.cell(row=row, column=5+pi*2, value=par)
            pc.fill = f(par_bg); pc.font = ft(bold=(par!="None"), size=8)
            pc.alignment = al("center","center"); pc.border = TB

        vc = ws.cell(row=row, column=MATRIX_COLS, value=cap.get("lastVerified",""))
        vc.fill = f(bg); vc.font = ft(size=8, color="6B7280")
        vc.alignment = al("center","center"); vc.border = TB
        ws.row_dimensions[row].height = 28
        row += 1

    for col, w in {"A":16,"B":26,"C":40}.items():
        ws.column_dimensions[col].width = w
    for pi, _ in enumerate(PROVIDERS):
        ws.column_dimensions[get_column_letter(4 + pi * 2)].width = 14
        ws.column_dimensions[get_column_letter(5 + pi * 2)].width = 16
    ws.column_dimensions[get_column_letter(MATRIX_COLS)].width = 12

def build_patterns_sheet(ws):
    """Architecture planning overlays grounded in the existing capability rows."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"
    pattern_cols = 6 + len(PROVIDERS)
    hdr(ws, "Architecture Patterns - Provider Framework-Informed Planning Overlays", pattern_cols)
    headers = [
        "Pattern / Rationale", "When To Use", "Capability",
    ]
    headers.extend(f"{PROV_LABELS[pkey]} Service / Gov" for pkey in PROVIDERS)
    headers.extend(["Review Questions", "Verification Boundary", "Verified"])
    for ci, header in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=header)
        c.fill = f("1E3A5F")
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center")
        c.border = TB
    ws.row_dimensions[2].height = 18

    row = 3
    for pattern in PATTERNS:
        for index, capability_name in enumerate(pattern["capabilities"]):
            cap = CAP_MAP[capability_name]
            is_first = index == 0
            values = [
                f"{pattern['name']}\n{pattern['summary']}" if is_first else "",
                pattern["whenToUse"] if is_first else "",
                capability_name,
            ]
            values.extend(
                f"{cap['providers'][pkey]['service']}\nGov: {cap['providers'][pkey]['govAvailability']} / Lag: {cap['providers'][pkey]['parityLag']}"
                for pkey in PROVIDERS
            )
            values.extend([
                "\n".join(pattern["reviewPrompts"]) if is_first else "",
                pattern["verificationNote"] if is_first else "",
                pattern["lastVerified"] if is_first else "",
            ])
            for ci, value in enumerate(values, 1):
                c = ws.cell(row=row, column=ci, value=value)
                c.fill = f("F8FAFC" if is_first else "FFFFFF")
                c.font = ft(bold=(ci == 3), size=8)
                c.alignment = al("left", "top", wrap=True)
                c.border = TB
            ws.row_dimensions[row].height = 68 if is_first else 34
            row += 1
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=pattern_cols)
    c = ws.cell(row=row, column=1, value="Official framework and enterprise foundation guidance")
    c.fill = f("0F1A2E")
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.alignment = al("left", "center")
    row += 1
    source_headers = ["Provider", "Architecture Framework", "Framework URL", "Enterprise Foundation Guidance", "Foundation URL", "Verified"]
    for ci, header in enumerate(source_headers, 1):
        c = ws.cell(row=row, column=ci, value=header)
        c.fill = f("1E3A5F")
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center")
        c.border = TB
    row += 1
    for pkey in PROVIDERS:
        guidance = FRAMEWORKS[pkey]
        values = [
            PROV_LABELS[pkey], guidance["framework"], guidance["frameworkUrl"],
            guidance["foundation"], guidance["foundationUrl"], guidance["lastVerified"],
        ]
        for ci, value in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=value)
            c.fill = f(PROV_COLORS[pkey]["svc"])
            c.font = ft(size=8, color="2563EB" if ci in [3, 5] else "111827")
            c.alignment = al("left", "center", wrap=True)
            c.border = TB
            if ci in [3, 5]:
                c.hyperlink = value
        row += 1

    for col, width in {"A":38, "B":38, "C":32}.items():
        ws.column_dimensions[col].width = width
    for index, _ in enumerate(PROVIDERS, 4):
        ws.column_dimensions[get_column_letter(index)].width = 35
    ws.column_dimensions[get_column_letter(4 + len(PROVIDERS))].width = 38
    ws.column_dimensions[get_column_letter(5 + len(PROVIDERS))].width = 40
    ws.column_dimensions[get_column_letter(6 + len(PROVIDERS))].width = 14

def build_compliance_sheet(ws):
    """Compliance framework references plus selected NIST SP 800-53 control-family planning rows."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    hdr(ws, "Compliance - Frameworks and NIST 800-53 Planning Lens", 9)
    headers = ["Framework", "Kind", "Issuer", "Status", "Scope", "NIST Alignment", "Historical Note", "Official Source", "Verified"]
    for ci, header in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=header)
        c.fill = f("1E3A5F")
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center")
        c.border = TB
    ws.row_dimensions[2].height = 18

    row = 3
    for framework in COMPLIANCE_FRAMEWORKS:
        values = [
            framework.get("name", ""),
            framework.get("kind", ""),
            framework.get("issuer", ""),
            framework.get("status", ""),
            framework.get("scope", ""),
            cell_text(framework.get("nistAlignment", "")),
            framework.get("historicalNote", ""),
            framework.get("url", ""),
            framework.get("lastVerified", ""),
        ]
        bg = "EFF6FF" if row % 2 else "F8FAFC"
        for ci, value in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=value)
            c.fill = f(bg)
            c.font = ft(bold=(ci == 1), size=8, color="2563EB" if ci == 8 and value else "111827")
            c.alignment = al("left", "top", wrap=True)
            c.border = TB
            if ci == 8 and value:
                c.hyperlink = value
        ws.row_dimensions[row].height = 58
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value=f"{CONTROL_LENS['name']} - {CONTROL_LENS['release']} control-family planning lens")
    c.fill = f("0F1A2E")
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.alignment = al("left", "center")
    row += 1

    family_headers = ["Family", "Implementation Focus", "Linked Capabilities", "Architecture Review Questions", "Boundary", "Catalog", "Baselines", "OSCAL", "Verified"]
    for ci, header in enumerate(family_headers, 1):
        c = ws.cell(row=row, column=ci, value=header)
        c.fill = f("1E3A5F")
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center")
        c.border = TB
    row += 1

    for index, family in enumerate(CONTROL_LENS["families"]):
        values = [
            f"{family['id']} - {family['name']}",
            family["applicability"],
            "\n".join(family["capabilities"]),
            "\n".join(family["reviewPrompts"]),
            CONTROL_LENS["scopeNote"] if index == 0 else "",
            CONTROL_LENS["catalogUrl"] if index == 0 else "",
            CONTROL_LENS["baselineUrl"] if index == 0 else "",
            CONTROL_LENS["oscalUrl"] if index == 0 else "",
            CONTROL_LENS["lastVerified"],
        ]
        bg = "ECFDF5" if index % 2 == 0 else "F8FAFC"
        for ci, value in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=value)
            c.fill = f(bg)
            c.font = ft(bold=(ci == 1), size=8, color="2563EB" if ci in [6, 7, 8] and value else "111827")
            c.alignment = al("left", "top", wrap=True)
            c.border = TB
            if ci in [6, 7, 8] and value:
                c.hyperlink = value
        ws.row_dimensions[row].height = 62
        row += 1

    for col, width in {"A":34, "B":24, "C":34, "D":18, "E":58, "F":58, "G":48, "H":54, "I":14}.items():
        ws.column_dimensions[col].width = width

def build_history_sheet(ws):
    """Provider journey milestones from official public sources."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    hdr(ws, "Cloud Provider History - Commercial, Free, and Government Milestones", 9)
    headers = ["Provider", "Phase", "Year", "Date", "Milestone", "Scope", "Summary", "Source", "Verified"]
    for ci, header in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=header)
        c.fill = f("1E3A5F")
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center")
        c.border = TB
    ws.row_dimensions[2].height = 18

    phase_bg = {
        "Commercial cloud": "DBEAFE",
        "Personal / Free": "D1FAE5",
        "Government state/federal": "FEF3C7",
    }
    sorted_history = sorted(HISTORY, key=lambda item: (item.get("year", 0), item.get("provider", ""), item.get("date", "")))
    for ri, item in enumerate(sorted_history, 3):
        pkey = item.get("provider", "")
        bg = phase_bg.get(item.get("phase", ""), PROV_COLORS.get(pkey, {}).get("svc", "F9FAFB"))
        vals = [
            PROV_LABELS.get(pkey, pkey.upper()),
            item.get("phase", ""),
            item.get("year", ""),
            item.get("dateLabel", item.get("date", "")),
            item.get("title", ""),
            " / ".join(item.get("scope", [])),
            item.get("summary", ""),
            item.get("sourceUrl", ""),
            hdata.get("_meta", {}).get("lastVerified", ""),
        ]
        for ci, value in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = f(bg)
            c.font = ft(bold=(ci in [1, 5]), size=8, color="2563EB" if ci == 8 else "111827")
            c.alignment = al("left", "top" if ci in [5, 7] else "center", wrap=True)
            c.border = TB
            if ci == 8 and value:
                c.hyperlink = value
        ws.row_dimensions[ri].height = 48

    note_row = len(sorted_history) + 5
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
    c = ws.cell(row=note_row, column=1, value=hdata.get("_meta", {}).get("scopeNote", ""))
    c.fill = f("F8FAFC")
    c.font = ft(size=8, italic=True, color="374151")
    c.alignment = al("left", "center", wrap=True)
    c.border = TB

    for col, width in {"A":10, "B":22, "C":8, "D":12, "E":38, "F":28, "G":66, "H":58, "I":12}.items():
        ws.column_dimensions[col].width = width

def build_transparency_sheet(ws):
    """State AI governance and transparency public-record scaffold."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    hdr(ws, "State AI Transparency - Official State Sources", 9)

    federal = TRANSPARENCY_META.get("federalContext", {})
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    c = ws.cell(row=2, column=1, value=f"Volatility note: {federal.get('citation', '')} - {federal.get('summary', '')}")
    c.fill = f("FEF3C7")
    c.font = ft(size=8, color="78350F", italic=True)
    c.alignment = al("left", "center", wrap=True)
    c.border = TB
    ws.row_dimensions[2].height = 34

    headers = ["State", "Status", "Instrument", "Title", "Citation", "Summary", "Official Source", "Verified", "Federal Context"]
    for ci, header in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=header)
        c.fill = f("1E3A5F")
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center")
        c.border = TB
    ws.row_dimensions[3].height = 18

    status_bg = {
        "Active": "D1FAE5",
        "Proposed": "FEF3C7",
        "Repealed": "FEE2E2",
        "None on record": "E5E7EB",
        "Unknown": "F8FAFC",
    }
    for ri, item in enumerate(TRANSPARENCY, 4):
        bg = status_bg.get(item.get("status", ""), "F8FAFC")
        vals = [
            f"{item.get('state', '')} - {item.get('stateName', '')}",
            item.get("status", ""),
            item.get("instrument", ""),
            item.get("title", ""),
            item.get("citation", ""),
            item.get("summary", ""),
            item.get("url", ""),
            item.get("lastVerified", ""),
            federal.get("url", "") if ri == 4 else "",
        ]
        for ci, value in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = f(bg)
            c.font = ft(bold=(ci in [1, 2]), size=8, color="2563EB" if ci in [7, 9] and value else "111827")
            c.alignment = al("left", "top" if ci in [4, 5, 6] else "center", wrap=True)
            c.border = TB
            if ci in [7, 9] and value:
                c.hyperlink = value
        ws.row_dimensions[ri].height = 42

    note_row = len(TRANSPARENCY) + 6
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
    c = ws.cell(row=note_row, column=1, value=TRANSPARENCY_META.get("scopeNote", ""))
    c.fill = f("F8FAFC")
    c.font = ft(size=8, italic=True, color="374151")
    c.alignment = al("left", "center", wrap=True)
    c.border = TB

    for col, width in {"A":22, "B":14, "C":20, "D":36, "E":36, "F":70, "G":54, "H":12, "I":54}.items():
        ws.column_dimensions[col].width = width

def build_ai_governance_sheet(ws, title, records, meta, accent="1E3A5F"):
    """Federal and international AI governance records from official public sources."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    hdr(ws, title, len(AI_GOVERNANCE_EXPORT_HEADERS))

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(AI_GOVERNANCE_EXPORT_HEADERS))
    c = ws.cell(row=2, column=1, value=meta.get("scopeNote", ""))
    c.fill = f("F8FAFC")
    c.font = ft(size=8, italic=True, color="374151")
    c.alignment = al("left", "center", wrap=True)
    c.border = TB
    ws.row_dimensions[2].height = 34

    headers = ["Jurisdiction", "Region", "Instrument", "Title", "Citation", "Status", "Summary", "Official Source", "Status / Signatures", "Verified", "Notes"]
    for ci, header in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=header)
        c.fill = f(accent)
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center")
        c.border = TB
    ws.row_dimensions[3].height = 18

    status_bg = {
        "Active": "D1FAE5",
        "Reference": "DBEAFE",
        "Revoked": "FEE2E2",
        "Superseded": "FEE2E2",
        "Binding treaty": "D1FAE5",
        "Regional law": "DBEAFE",
        "Non-binding framework": "FEF3C7",
        "Resolution": "E5E7EB",
        "Voluntary framework": "FEF3C7",
        "Placeholder": "F8FAFC",
    }
    for ri, item in enumerate(records, 4):
        bg = status_bg.get(item.get("status", ""), "F8FAFC")
        vals = [
            item.get("jurisdiction", ""),
            item.get("region", ""),
            item.get("instrument", ""),
            item.get("title", ""),
            item.get("citation", ""),
            item.get("status", ""),
            item.get("summary", ""),
            item.get("url", ""),
            item.get("statusUrl", ""),
            item.get("lastVerified", ""),
            item.get("notes", ""),
        ]
        for ci, value in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = f(bg)
            c.font = ft(bold=(ci in [1, 6]), size=8, color="2563EB" if ci in [8, 9] and value else "111827")
            c.alignment = al("left", "top" if ci in [4, 5, 7, 11] else "center", wrap=True)
            c.border = TB
            if ci in [8, 9] and value:
                c.hyperlink = value
        ws.row_dimensions[ri].height = 46

    for col, width in {"A":28,"B":18,"C":24,"D":46,"E":38,"F":20,"G":72,"H":56,"I":56,"J":13,"K":42}.items():
        ws.column_dimensions[col].width = width

def build_upcoming_sheet(ws):
    ws.sheet_view.showGridLines = False
    hdr(ws, "Announced / Preview / Upcoming — Official Sources Only", 10)
    hdrs = ["ID","Provider","Category","Type","Status","Title","Announced","Expected GA","Source","Detail","Verified"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = f("78350F"); c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center","center"); c.border = TB
    ws.row_dimensions[2].height = 16

    STATUS_BG = {"ga":"D1FAE5","preview":"EDE9FE","announced":"FEF3C7","limited":"FFE4E6"}
    for ri, item in enumerate(UPCOMING, 3):
        bg = STATUS_BG.get(item.get("status",""),"F9FAFB")
        vals = [item.get("id",""),item.get("provider",""),item.get("category",""),
                item.get("type",""),item.get("status",""),item.get("title",""),
                item.get("announced",""),item.get("expected_ga",""),item.get("source",""),item.get("detail",""),
                str(item.get("verified",""))]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = f(bg); c.font = ft(size=8)
            c.alignment = al("left","center", wrap=True); c.border = TB
        ws.row_dimensions[ri].height = 34

    for col, w in {"A":22,"B":8,"C":22,"D":18,"E":12,"F":38,"G":14,"H":12,"I":50,"J":60,"K":10}.items():
        ws.column_dimensions[col].width = w

# ── Build ──────────────────────────────────────────────────────────────────
def build_status_sheet(ws):
    ws.sheet_view.showGridLines = False
    hdr(ws, "Operational Status Sources - Official Provider Pages", 8)
    hdrs = ["Provider","Category","Name","Summary","Status URL","History URL","Docs URL","Verified"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = f("1E3A5F"); c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center","center"); c.border = TB
    ws.row_dimensions[2].height = 16

    for ri, item in enumerate(STATUS_SOURCES, 3):
        bg = "F8FAFC" if item.get("category") == "cloud-provider" else "FFFBEB"
        vals = [
            item.get("providerName", ""),
            item.get("category", ""),
            item.get("name", ""),
            item.get("summary", ""),
            item.get("statusUrl", ""),
            item.get("historyUrl", ""),
            item.get("docsUrl", ""),
            item.get("lastVerified", ""),
        ]
        for ci, value in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = f(bg); c.font = ft(size=8, color="2563EB" if ci in [5, 6, 7] and value else "111827")
            c.alignment = al("left","center", wrap=True); c.border = TB
            if ci in [5, 6, 7] and value:
                c.hyperlink = value
        ws.row_dimensions[ri].height = 34

    for col, w in {"A":14,"B":18,"C":28,"D":64,"E":48,"F":48,"G":48,"H":12}.items():
        ws.column_dimensions[col].width = w

def ai_watch_model_rows():
    rows = []
    for item in sorted(AI_WATCH_SOURCES, key=lambda entry: (entry.get("category", ""), entry.get("name", ""))):
        details = item.get("modelDetails", [])
        if not isinstance(details, list) or not details:
            details = [{
                "name": "; ".join(item.get("models", [])),
                "bestFor": item.get("summary", ""),
                "sourceNote": "",
                "releaseDate": "",
                "docUrl": item.get("docsUrl", ""),
                "releaseNotesUrl": item.get("releaseNotesUrl", ""),
                "lastVerified": item.get("lastVerified", ""),
            }]
        for detail in details:
            if not isinstance(detail, dict):
                continue
            rows.append({
                "lab": item.get("name", ""),
                "shortName": item.get("shortName", ""),
                "category": item.get("category", ""),
                "modelFamily": item.get("modelFamily", ""),
                "modelName": detail.get("name", ""),
                "bestFor": detail.get("bestFor", ""),
                "sourceNote": detail.get("sourceNote", ""),
                "releaseDate": detail.get("releaseDate", ""),
                "modelDocUrl": detail.get("docUrl", ""),
                "modelReleaseNotesUrl": detail.get("releaseNotesUrl", ""),
                "modelLastVerified": detail.get("lastVerified", ""),
                "labNewsUrl": item.get("newsUrl", ""),
                "labDocsUrl": item.get("docsUrl", ""),
                "safetyUrl": item.get("safetyUrl", ""),
                "labLastVerified": item.get("lastVerified", ""),
            })
    return rows

def build_ai_watch_sheet(ws):
    ws.sheet_view.showGridLines = False
    hdr(ws, "Foundational & Frontier Releases - Official Model Selection Notes", 15)
    hdrs = ["Lab","Short Name","Category","Model Family","Model","Best For","Source Note","Release Date","Model Docs URL","Model Release Notes URL","Model Verified","Lab News URL","Lab Docs URL","Safety URL","Lab Verified"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = f("0B3B66"); c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center","center"); c.border = TB
    ws.row_dimensions[2].height = 16

    for ri, item in enumerate(ai_watch_model_rows(), 3):
        bg = {
            "frontier-model-lab": "EFF6FF",
            "open-model-lab": "ECFDF5",
            "multimodal-model-lab": "FFFBEB",
        }.get(item.get("category"), "F8FAFC")
        vals = [
            item.get("lab", ""),
            item.get("shortName", ""),
            item.get("category", ""),
            item.get("modelFamily", ""),
            item.get("modelName", ""),
            item.get("bestFor", ""),
            item.get("sourceNote", ""),
            item.get("releaseDate", ""),
            item.get("modelDocUrl", ""),
            item.get("modelReleaseNotesUrl", ""),
            item.get("modelLastVerified", ""),
            item.get("labNewsUrl", ""),
            item.get("labDocsUrl", ""),
            item.get("safetyUrl", ""),
            item.get("labLastVerified", ""),
        ]
        for ci, value in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = f(bg); c.font = ft(size=8, color="2563EB" if ci in [9, 10, 12, 13, 14] and value else "111827")
            c.alignment = al("left","center", wrap=True); c.border = TB
            if ci in [9, 10, 12, 13, 14] and value:
                c.hyperlink = value
        ws.row_dimensions[ri].height = 36

    for col, w in {"A":18,"B":14,"C":22,"D":28,"E":24,"F":52,"G":70,"H":13,"I":46,"J":46,"K":12,"L":42,"M":46,"N":42,"O":12}.items():
        ws.column_dimensions[col].width = w

MATRIX_EXPORT_HEADERS = [
    "capability", "category", "tags", "aiClassification", "provider", "service",
    "formerNames", "status", "govAvailability", "parityLag", "parityDetail", "govVariant", "region", "realmClass",
    "providerLastVerified", "constraints", "costShape", "egressSensitive", "commitmentDiscountAvailable",
    "pqcStatus", "pqcKem", "pqcSignature", "pqcTls", "pqcVpn", "pqcMilestoneDate", "pqcFipsEndpointParity",
    "pqcGovPqc", "pqcSource", "pqcSourceDate", "pqcFirstParty", "pqcConfidence", "pqcNote",
    "residencyOfferings", "residencyGeographies", "residencyStatuses", "residencyPartnerOperated", "residencySources",
    "fedrampCommercialStatus", "fedrampCommercialUrl", "fedrampCommercialDate", "fedrampCommercialConfidence",
    "fedrampGovernmentStatus", "fedrampGovernmentDodIL", "fedrampGovernmentBoundary", "fedrampGovernmentUrl",
    "fedrampGovernmentDate", "fedrampGovernmentConfidence", "fedrampLevel", "dodImpactLevel", "docsUrl",
    "govDocsUrl", "complianceUrl", "pricingUrl", "lastVerified", "sourceNotes",
]
PATTERN_EXPORT_HEADERS = [
    "pattern", "summary", "whenToUse", "capability", "category", "provider",
    "service", "govAvailability", "parityLag", "providerFramework", "frameworkUrl",
    "providerFoundation", "foundationUrl", "reviewPrompts", "verificationNote", "lastVerified",
]
COMPLIANCE_EXPORT_HEADERS = [
    "rowType", "id", "name", "kind", "issuer", "status", "scope", "nistAlignment",
    "historicalNote", "officialUrl", "linkedCapabilities", "reviewPrompts", "lastVerified",
]
HISTORY_EXPORT_HEADERS = [
    "provider", "phase", "year", "date", "dateLabel", "title", "summary",
    "scope", "sourceLabel", "sourceUrl", "lastVerified",
]
TRANSPARENCY_EXPORT_HEADERS = [
    "state", "stateName", "instrument", "title", "citation", "status",
    "summary", "url", "lastVerified",
]
AI_GOVERNANCE_EXPORT_HEADERS = [
    "jurisdiction", "region", "instrument", "title", "citation", "status",
    "summary", "url", "statusUrl", "lastVerified", "notes",
]
STATUS_EXPORT_HEADERS = [
    "providerName", "category", "name", "summary", "statusUrl", "historyUrl",
    "docsUrl", "lastVerified",
]
AI_WATCH_EXPORT_HEADERS = [
    "lab", "shortName", "category", "modelFamily", "modelName", "bestFor",
    "sourceNote", "releaseDate", "modelDocUrl", "modelReleaseNotesUrl",
    "modelLastVerified", "labNewsUrl", "labDocsUrl", "safetyUrl", "labLastVerified",
]
RESIDENCY_EXPORT_HEADERS = [
    "capability", "provider", "offering", "guarantee", "geography", "status",
    "firstParty", "operatingModel", "source", "providerVerified", "matrixVersion",
]

def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)
    return str(value)

def field_value(value, field):
    if isinstance(value, dict):
        return value.get(field, "")
    return ""

def residency_items(value):
    return value if isinstance(value, list) else []

def residency_field(value, field):
    return "; ".join(
        str(item.get(field, ""))
        for item in residency_items(value)
        if item.get(field, "") not in [None, ""]
    )

def residency_partner_operated(value):
    return "; ".join(
        str(item.get("offering", ""))
        for item in residency_items(value)
        if item.get("firstParty") is False and item.get("offering")
    )

def export_notes(*parts):
    return " ".join(cell_text(part).strip() for part in parts if cell_text(part).strip())

def matrix_export_rows(caps):
    rows = []
    for cap in caps:
        for pkey in PROVIDERS:
            provider = cap.get("providers", {}).get(pkey, {})
            cost_model = provider.get("costModel", {})
            pqc = provider.get("pqcReadiness", {})
            residency = provider.get("residency", [])
            fedramp = provider.get("fedramp", {})
            fedramp_commercial = field_value(fedramp, "commercial") if isinstance(field_value(fedramp, "commercial"), dict) else {}
            fedramp_government = field_value(fedramp, "government") if isinstance(field_value(fedramp, "government"), dict) else {}
            rows.append({
                "capability": cap.get("capability", ""),
                "category": cap.get("category", ""),
                "tags": cap.get("tags", []),
                "aiClassification": cap.get("aiClassification", ""),
                "provider": PROV_LABELS.get(pkey, pkey.upper()),
                "service": provider.get("service", ""),
                "formerNames": provider.get("formerNames", []),
                "status": provider.get("status", ""),
                "govAvailability": provider.get("govAvailability", ""),
                "parityLag": provider.get("parityLag", ""),
                "parityDetail": provider.get("parityDetail", ""),
                "govVariant": provider.get("govVariant", ""),
                "region": provider.get("region", ""),
                "realmClass": provider.get("realmClass", ""),
                "providerLastVerified": provider.get("lastVerified", ""),
                "constraints": provider.get("constraints", ""),
                "costShape": field_value(cost_model, "shape"),
                "egressSensitive": field_value(cost_model, "egressSensitive"),
                "commitmentDiscountAvailable": field_value(cost_model, "commitmentDiscountAvailable"),
                "pqcStatus": field_value(pqc, "status"),
                "pqcKem": field_value(pqc, "kem"),
                "pqcSignature": field_value(pqc, "signature"),
                "pqcTls": field_value(pqc, "tls"),
                "pqcVpn": field_value(pqc, "vpn"),
                "pqcMilestoneDate": field_value(pqc, "milestoneDate"),
                "pqcFipsEndpointParity": field_value(pqc, "fipsEndpointParity"),
                "pqcGovPqc": field_value(pqc, "govPqc"),
                "pqcSource": field_value(pqc, "source"),
                "pqcSourceDate": field_value(pqc, "sourceDate"),
                "pqcFirstParty": field_value(pqc, "firstParty"),
                "pqcConfidence": field_value(pqc, "confidence"),
                "pqcNote": field_value(pqc, "note"),
                "residencyOfferings": residency_field(residency, "offering"),
                "residencyGeographies": residency_field(residency, "geography"),
                "residencyStatuses": residency_field(residency, "status"),
                "residencyPartnerOperated": residency_partner_operated(residency),
                "residencySources": residency_field(residency, "source"),
                "fedrampCommercialStatus": field_value(fedramp_commercial, "status"),
                "fedrampCommercialUrl": field_value(fedramp_commercial, "url"),
                "fedrampCommercialDate": field_value(fedramp_commercial, "date"),
                "fedrampCommercialConfidence": field_value(fedramp_commercial, "confidence"),
                "fedrampGovernmentStatus": field_value(fedramp_government, "status"),
                "fedrampGovernmentDodIL": field_value(fedramp_government, "dodIL"),
                "fedrampGovernmentBoundary": field_value(fedramp_government, "boundary"),
                "fedrampGovernmentUrl": field_value(fedramp_government, "url"),
                "fedrampGovernmentDate": field_value(fedramp_government, "date"),
                "fedrampGovernmentConfidence": field_value(fedramp_government, "confidence"),
                "fedrampLevel": provider.get("fedrampLevel", ""),
                "dodImpactLevel": provider.get("dodImpactLevel", ""),
                "docsUrl": provider.get("docsUrl", ""),
                "govDocsUrl": provider.get("govDocsUrl", ""),
                "complianceUrl": provider.get("complianceUrl", ""),
                "pricingUrl": provider.get("pricingUrl", ""),
                "lastVerified": cap.get("lastVerified", ""),
                "sourceNotes": export_notes(cap.get("sourceNotes", ""), provider.get("sourceNotes", "")),
            })
    return rows

def residency_export_rows():
    rows = []
    for cap in CAPS:
        for pkey in PROVIDERS:
            provider = cap.get("providers", {}).get(pkey, {})
            for item in residency_items(provider.get("residency", [])):
                rows.append({
                    "capability": cap.get("capability", ""),
                    "provider": PROV_LABELS.get(pkey, pkey.upper()),
                    "offering": item.get("offering", ""),
                    "guarantee": item.get("guarantee", ""),
                    "geography": item.get("geography", ""),
                    "status": item.get("status", ""),
                    "firstParty": item.get("firstParty", ""),
                    "operatingModel": "First-party" if item.get("firstParty") is True else "Partner-operated",
                    "source": item.get("source", ""),
                    "providerVerified": provider.get("lastVerified", ""),
                    "matrixVersion": META.get("version", ""),
                })
    return rows

def build_residency_sheet(ws):
    ws.sheet_view.showGridLines = False
    rows = residency_export_rows()
    hdr(ws, "Residency and sovereignty offerings", len(RESIDENCY_EXPORT_HEADERS))
    for ci, header in enumerate(RESIDENCY_EXPORT_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=header)
        c.fill = f("1E3A5F"); c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.alignment = al("center", "center"); c.border = TB
    for ri, row_data in enumerate(rows, 3):
        provider_name = row_data.get("provider", "")
        pkey = next((key for key, label in PROV_LABELS.items() if label == provider_name), "aws")
        bg = PROV_COLORS[pkey]["svc"]
        for ci, header in enumerate(RESIDENCY_EXPORT_HEADERS, 1):
            value = row_data.get(header, "")
            c = ws.cell(row=ri, column=ci, value=value)
            c.fill = f(bg); c.font = ft(size=8, color="2563EB" if header == "source" and value else "111827")
            c.alignment = al("left", "center", wrap=True); c.border = TB
            if header == "source" and value:
                c.hyperlink = value
        ws.row_dimensions[ri].height = 36
    for col, width in {"A":32,"B":10,"C":30,"D":58,"E":34,"F":14,"G":12,"H":18,"I":54,"J":14,"K":12}.items():
        ws.column_dimensions[col].width = width

def pattern_export_rows():
    rows = []
    for pattern in PATTERNS:
        for capability_name in pattern.get("capabilities", []):
            cap = CAP_MAP.get(capability_name)
            if not cap:
                continue
            for pkey in PROVIDERS:
                provider = cap.get("providers", {}).get(pkey, {})
                framework = FRAMEWORKS.get(pkey, {})
                rows.append({
                    "pattern": pattern.get("name", ""),
                    "summary": pattern.get("summary", ""),
                    "whenToUse": pattern.get("whenToUse", ""),
                    "capability": capability_name,
                    "category": cap.get("category", ""),
                    "provider": PROV_LABELS.get(pkey, pkey.upper()),
                    "service": provider.get("service", ""),
                    "govAvailability": provider.get("govAvailability", ""),
                    "parityLag": provider.get("parityLag", ""),
                    "providerFramework": framework.get("framework", ""),
                    "frameworkUrl": framework.get("frameworkUrl", ""),
                    "providerFoundation": framework.get("foundation", ""),
                    "foundationUrl": framework.get("foundationUrl", ""),
                    "reviewPrompts": pattern.get("reviewPrompts", []),
                    "verificationNote": pattern.get("verificationNote", ""),
                    "lastVerified": pattern.get("lastVerified", ""),
                })
    return rows

def compliance_export_rows():
    rows = [
        {
            "rowType": "Framework",
            "id": framework.get("id", ""),
            "name": framework.get("name", ""),
            "kind": framework.get("kind", ""),
            "issuer": framework.get("issuer", ""),
            "status": framework.get("status", ""),
            "scope": framework.get("scope", ""),
            "nistAlignment": framework.get("nistAlignment", ""),
            "historicalNote": framework.get("historicalNote", ""),
            "officialUrl": framework.get("url", ""),
            "linkedCapabilities": "",
            "reviewPrompts": "",
            "lastVerified": framework.get("lastVerified", ""),
        }
        for framework in COMPLIANCE_FRAMEWORKS
    ]
    rows.extend(
        {
            "rowType": "NIST control family",
            "id": family.get("id", ""),
            "name": family.get("name", ""),
            "kind": CONTROL_LENS.get("id", ""),
            "issuer": "NIST",
            "status": CONTROL_LENS.get("release", ""),
            "scope": family.get("applicability", ""),
            "nistAlignment": CONTROL_LENS.get("scopeNote", ""),
            "historicalNote": "",
            "officialUrl": CONTROL_LENS.get("catalogUrl", ""),
            "linkedCapabilities": family.get("capabilities", []),
            "reviewPrompts": family.get("reviewPrompts", []),
            "lastVerified": CONTROL_LENS.get("lastVerified", ""),
        }
        for family in CONTROL_LENS.get("families", [])
    )
    return rows

def history_export_rows():
    return [
        {
            "provider": PROV_LABELS.get(item.get("provider", ""), item.get("provider", "").upper()),
            "phase": item.get("phase", ""),
            "year": item.get("year", ""),
            "date": item.get("date", ""),
            "dateLabel": item.get("dateLabel", ""),
            "title": item.get("title", ""),
            "summary": item.get("summary", ""),
            "scope": item.get("scope", []),
            "sourceLabel": item.get("sourceLabel", ""),
            "sourceUrl": item.get("sourceUrl", ""),
            "lastVerified": hdata.get("_meta", {}).get("lastVerified", ""),
        }
        for item in sorted(HISTORY, key=lambda entry: (entry.get("year", 0), entry.get("provider", ""), entry.get("date", "")))
    ]

def transparency_export_rows():
    return [
        {
            "state": item.get("state", ""),
            "stateName": item.get("stateName", ""),
            "instrument": item.get("instrument", ""),
            "title": item.get("title", ""),
            "citation": item.get("citation", ""),
            "status": item.get("status", ""),
            "summary": item.get("summary", ""),
            "url": item.get("url", ""),
            "lastVerified": item.get("lastVerified", ""),
        }
        for item in sorted(TRANSPARENCY, key=lambda entry: (entry.get("stateName", ""), entry.get("title", "")))
    ]

def ai_governance_export_rows(records):
    return [
        {
            "jurisdiction": item.get("jurisdiction", ""),
            "region": item.get("region", ""),
            "instrument": item.get("instrument", ""),
            "title": item.get("title", ""),
            "citation": item.get("citation", ""),
            "status": item.get("status", ""),
            "summary": item.get("summary", ""),
            "url": item.get("url", ""),
            "statusUrl": item.get("statusUrl", ""),
            "lastVerified": item.get("lastVerified", ""),
            "notes": item.get("notes", ""),
        }
        for item in sorted(records, key=lambda entry: (entry.get("jurisdiction", ""), entry.get("title", "")))
    ]

def status_export_rows():
    return [
        {
            "providerName": item.get("providerName", ""),
            "category": item.get("category", ""),
            "name": item.get("name", ""),
            "summary": item.get("summary", ""),
            "statusUrl": item.get("statusUrl", ""),
            "historyUrl": item.get("historyUrl", ""),
            "docsUrl": item.get("docsUrl", ""),
            "lastVerified": item.get("lastVerified", ""),
        }
        for item in sorted(STATUS_SOURCES, key=lambda entry: (entry.get("category", ""), entry.get("providerName", "")))
    ]

def ai_watch_export_rows():
    return ai_watch_model_rows()

def write_view_csv(view_id, headers, rows):
    out = OUTDIR / f"cloudintelmatrix-{view_id}-{EXPORT_DATE}.csv"
    stable_out = OUTDIR / f"cloudintelmatrix-{view_id}.csv"
    for target in [out, stable_out]:
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({header: cell_text(row.get(header, "")) for header in headers})
    print(f"Saved CSV: {out} ({len(rows)} row(s))")

def write_view_csvs():
    matrix_rows = matrix_export_rows(CAPS)
    write_view_csv("matrix", MATRIX_EXPORT_HEADERS, matrix_rows)
    write_view_csv("diff", MATRIX_EXPORT_HEADERS, matrix_rows)
    write_view_csv("gov", MATRIX_EXPORT_HEADERS, matrix_rows)
    write_view_csv(
        "ai",
        MATRIX_EXPORT_HEADERS,
        matrix_export_rows([cap for cap in CAPS if any(tag in {"AI_NATIVE", "AI_CAPABLE"} for tag in cap.get("tags", []))]),
    )
    write_view_csv("patterns", PATTERN_EXPORT_HEADERS, pattern_export_rows())
    write_view_csv("controls", COMPLIANCE_EXPORT_HEADERS, compliance_export_rows())
    write_view_csv("history", HISTORY_EXPORT_HEADERS, history_export_rows())
    write_view_csv("transparency", TRANSPARENCY_EXPORT_HEADERS, transparency_export_rows())
    write_view_csv("federal-transparency", AI_GOVERNANCE_EXPORT_HEADERS, ai_governance_export_rows(FEDERAL_TRANSPARENCY))
    write_view_csv("international-transparency", AI_GOVERNANCE_EXPORT_HEADERS, ai_governance_export_rows(INTERNATIONAL_TRANSPARENCY))
    write_view_csv("status", STATUS_EXPORT_HEADERS, status_export_rows())
    write_view_csv("ai-watch", AI_WATCH_EXPORT_HEADERS, ai_watch_export_rows())
    write_view_csv("residency", RESIDENCY_EXPORT_HEADERS, residency_export_rows())

wb = Workbook()
wb.remove(wb.active)

ws1 = wb.create_sheet("Matrix — All Tiers")
build_matrix_sheet(ws1)

for tier in TIERS:
    short = {"Personal / Free":"Free","Commercial / SMB":"SMB","Enterprise":"Enterprise","Government":"Govt"}[tier]
    ws = wb.create_sheet(f"Matrix — {short}")
    build_matrix_sheet(ws, tier=tier)

ws_detail = wb.create_sheet("Full Detail")
build_detail_sheet(ws_detail)

ws_gov = wb.create_sheet("Gov & Parity")
build_gov_sheet(ws_gov)

ws_residency = wb.create_sheet("Residency")
build_residency_sheet(ws_residency)

ws_patterns = wb.create_sheet("Architecture Patterns")
build_patterns_sheet(ws_patterns)

ws_controls = wb.create_sheet("Compliance")
build_compliance_sheet(ws_controls)

ws_history = wb.create_sheet("Cloud History")
build_history_sheet(ws_history)

ws_transparency = wb.create_sheet("Transparency")
build_transparency_sheet(ws_transparency)

ws_federal_ai = wb.create_sheet("Federal AI")
build_ai_governance_sheet(
    ws_federal_ai,
    "Federal AI Transparency - Official Federal Sources",
    FEDERAL_TRANSPARENCY,
    FEDERAL_TRANSPARENCY_META,
    accent="1D4ED8",
)

ws_international_ai = wb.create_sheet("International AI")
build_ai_governance_sheet(
    ws_international_ai,
    "International AI Transparency - Official Framework Sources",
    INTERNATIONAL_TRANSPARENCY,
    INTERNATIONAL_TRANSPARENCY_META,
    accent="0E7490",
)

ws_up = wb.create_sheet("Upcoming & Future")
build_upcoming_sheet(ws_up)

ws_status = wb.create_sheet("Status Sources")
build_status_sheet(ws_status)

ws_ai_watch = wb.create_sheet("Model Releases")
build_ai_watch_sheet(ws_ai_watch)

out = OUTDIR / "Cloud_Intelligence_Matrix.xlsx"
wb.save(out)
write_view_csvs()
print(f"✅ Saved: {out}  ({out.stat().st_size // 1024} KB)")
